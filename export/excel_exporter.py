from __future__ import annotations

"""
InfraAuto - Professional Excel Exporter
=========================================
openpyxl 기반 6-시트 Excel 워크북 생성 모듈.

시트 구성:
  1. 표지       - 프로젝트 정보 및 총원가 요약
  2. 견적서     - 항목별 견적 내역
  3. 세부공정내역 - 자재/노무 상세 공정
  4. 공정별요약   - 대공정 기준 집계
  5. 남은작업    - 잔여 공정 추적 (진행률, 상태)
  6. 대시보드    - 비용 분석 차트
"""

from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    ALT_ROW_COLOR,
    AUTHOR_NAME,
    COMPANY_NAME,
    HEADER_BG_COLOR,
    HEADER_FONT_COLOR,
    PROJECT_NAME,
    STATUS_COMPLETE_COLOR,
    STATUS_PROGRESS_COLOR,
    STATUS_WAITING_COLOR,
    TOTAL_ROW_BG_COLOR,
    VERSION,
)

# ---------------------------------------------------------------------------
# Reusable style constants
# ---------------------------------------------------------------------------

_FONT_NAME = "맑은 고딕"

_HEADER_FILL = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color=HEADER_FONT_COLOR, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ALT_ROW_FILL = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")

_TOTAL_FILL = PatternFill(start_color=TOTAL_ROW_BG_COLOR, end_color=TOTAL_ROW_BG_COLOR, fill_type="solid")
_TOTAL_FONT = Font(name=_FONT_NAME, bold=True, size=11)

_DATA_FONT = Font(name=_FONT_NAME, size=10)
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
_DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGNMENT_RIGHT = Alignment(horizontal="right", vertical="center")

_THIN_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_THICK_TOP_SIDE = Side(style="medium", color="000000")
_TOTAL_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THICK_TOP_SIDE,
    bottom=_THIN_SIDE,
)

_NUMBER_FORMAT = "#,##0"
_PERCENT_FORMAT = "0.0%"

_STATUS_FILLS = {
    "완료": PatternFill(start_color=STATUS_COMPLETE_COLOR, end_color=STATUS_COMPLETE_COLOR, fill_type="solid"),
    "진행중": PatternFill(start_color=STATUS_PROGRESS_COLOR, end_color=STATUS_PROGRESS_COLOR, fill_type="solid"),
    "대기": PatternFill(start_color=STATUS_WAITING_COLOR, end_color=STATUS_WAITING_COLOR, fill_type="solid"),
}


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _auto_width(ws: Worksheet, *, min_width: float = 8, max_width: float = 50) -> None:
    """Adjust column widths based on cell content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                text = str(cell.value)
                # Approximate: Korean characters count as ~2 width units
                char_len = sum(2 if ord(ch) > 127 else 1 for ch in text)
                # Handle multiline cells
                lines = text.split("\n")
                widest_line = max(
                    sum(2 if ord(ch) > 127 else 1 for ch in line) for line in lines
                )
                char_len = widest_line
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _apply_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Write header values and apply header styling to a row."""
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _apply_data_cell(
    cell: Any,
    *,
    is_number: bool = False,
    is_center: bool = False,
    row_idx: int = 0,
) -> None:
    """Apply standard data cell styling."""
    cell.font = _DATA_FONT
    cell.border = _THIN_BORDER

    if is_number:
        cell.number_format = _NUMBER_FORMAT
        cell.alignment = _DATA_ALIGNMENT_RIGHT
    elif is_center:
        cell.alignment = _DATA_ALIGNMENT_CENTER
    else:
        cell.alignment = _DATA_ALIGNMENT

    # Alternating row color (0-indexed: even data rows get fill)
    if row_idx % 2 == 0:
        cell.fill = _ALT_ROW_FILL


def _apply_total_row(ws: Worksheet, row: int, col_count: int) -> None:
    """Apply total-row styling across all columns in the given row."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.border = _TOTAL_BORDER


# ---------------------------------------------------------------------------
# ProfessionalExcelExporter
# ---------------------------------------------------------------------------

class ProfessionalExcelExporter:
    """Generate a professionally formatted 6-sheet Excel workbook for InfraAuto.

    Sheets:
        1. 표지 (Cover)
        2. 견적서 (Estimate)
        3. 세부공정내역 (Process Detail)
        4. 공정별요약 (Process Summary)
        5. 남은작업 (Remaining Work)
        6. 대시보드 (Dashboard)
    """

    def export(
        self,
        file_path: str,
        estimate_rows: list[dict],
        processes: list[dict],
        summary: dict,
        remaining_work: list[dict] | None = None,
        project_info: dict | None = None,
    ) -> str:
        """Create and save the Excel workbook.

        Args:
            file_path: Output .xlsx path.
            estimate_rows: List of dicts with keys:
                item_name, quantity, unit, unit_price, total, basis.
            processes: List of dicts with keys:
                no, process, sub_process, material, quantity, unit,
                material_cost, labor, labor_days, labor_cost, total_cost.
            summary: Dict of ``{process_name: {process, material_cost,
                labor_cost, total_cost, total_days, sub_count}}``.
            remaining_work: Optional list of dicts for pre-filled remaining
                work data.  Expected keys: process, sub_process, material,
                quantity, unit, completed_qty (default 0), status
                (default "대기"), note.
            project_info: Optional dict with keys ``company_name``,
                ``project_name``, ``author``.

        Returns:
            The *file_path* that was written.
        """
        info = self._resolve_project_info(project_info)
        grand_total = sum(r.get("total", 0) for r in estimate_rows)

        wb = Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        self._build_cover(wb, info, grand_total)
        self._build_estimate(wb, estimate_rows, grand_total)
        self._build_process_detail(wb, processes)
        self._build_process_summary(wb, summary)
        self._build_remaining_work(wb, processes, remaining_work)
        self._build_dashboard(wb, summary)

        wb.save(file_path)
        return file_path

    # ------------------------------------------------------------------
    # Project info resolution
    # ------------------------------------------------------------------

    @staticmethod
    def _resolve_project_info(project_info: dict | None) -> dict:
        defaults = {
            "company_name": COMPANY_NAME,
            "project_name": PROJECT_NAME,
            "author": AUTHOR_NAME,
        }
        if project_info:
            defaults.update({k: v for k, v in project_info.items() if v})
        return defaults

    # ------------------------------------------------------------------
    # Sheet 1: 표지 (Cover)
    # ------------------------------------------------------------------

    def _build_cover(self, wb: Workbook, info: dict, grand_total: int) -> None:
        ws = wb.create_sheet(title="표지")

        # Merge a wide area for the title block
        ws.merge_cells("B3:G3")
        ws.merge_cells("B5:G5")
        ws.merge_cells("B7:G7")
        ws.merge_cells("B9:G9")
        ws.merge_cells("B11:G11")
        ws.merge_cells("B13:G13")
        ws.merge_cells("B15:G15")

        # Title
        title_cell = ws["B3"]
        title_cell.value = "견 적 서"
        title_cell.font = Font(name=_FONT_NAME, bold=True, size=28, color=HEADER_BG_COLOR)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Subtitle / version
        sub_cell = ws["B5"]
        sub_cell.value = f"InfraAuto v{VERSION} - 건설 견적 자동화 시스템"
        sub_cell.font = Font(name=_FONT_NAME, size=12, color="666666")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Horizontal rule (using bottom border on a row)
        for col in range(2, 8):
            ws.cell(row=6, column=col).border = Border(
                bottom=Side(style="medium", color=HEADER_BG_COLOR)
            )

        # Project name
        ws["B7"].value = f"프로젝트: {info['project_name']}"
        ws["B7"].font = Font(name=_FONT_NAME, size=14, bold=True)
        ws["B7"].alignment = Alignment(horizontal="center", vertical="center")

        # Company
        ws["B9"].value = f"회사명: {info['company_name']}"
        ws["B9"].font = Font(name=_FONT_NAME, size=12)
        ws["B9"].alignment = Alignment(horizontal="center", vertical="center")

        # Author
        author_text = info["author"] if info["author"] else "-"
        ws["B11"].value = f"작성자: {author_text}"
        ws["B11"].font = Font(name=_FONT_NAME, size=12)
        ws["B11"].alignment = Alignment(horizontal="center", vertical="center")

        # Date
        ws["B13"].value = f"작성일: {date.today().strftime('%Y년 %m월 %d일')}"
        ws["B13"].font = Font(name=_FONT_NAME, size=12)
        ws["B13"].alignment = Alignment(horizontal="center", vertical="center")

        # Grand total
        ws["B15"].value = f"총 견적금액: {grand_total:,} 원"
        ws["B15"].font = Font(name=_FONT_NAME, size=16, bold=True, color=HEADER_BG_COLOR)
        ws["B15"].alignment = Alignment(horizontal="center", vertical="center")

        # Set row heights for visual spacing
        for r in range(1, 17):
            ws.row_dimensions[r].height = 30
        ws.row_dimensions[3].height = 55
        ws.row_dimensions[15].height = 45

        # Set column widths for the cover
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # ------------------------------------------------------------------
    # Sheet 2: 견적서 (Estimate)
    # ------------------------------------------------------------------

    def _build_estimate(self, wb: Workbook, estimate_rows: list[dict], grand_total: int) -> None:
        ws = wb.create_sheet(title="견적서")

        headers = ["No", "항목", "수량", "단위", "단가", "금액", "산출근거"]
        _apply_header_row(ws, row=1, headers=headers)

        number_cols = {3, 5, 6}  # quantity, unit_price, total (1-indexed)
        center_cols = {1, 4}     # No, unit

        for row_idx, item in enumerate(estimate_rows):
            r = row_idx + 2  # data starts at row 2
            values = [
                row_idx + 1,
                item.get("item_name", "") or item.get("name", ""),
                item.get("quantity", 0) or item.get("qty", 0),
                item.get("unit", ""),
                item.get("unit_price", 0),
                item.get("total", 0),
                str(item.get("basis", "")).replace("\n", " | "),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                _apply_data_cell(
                    cell,
                    is_number=(col_idx in number_cols),
                    is_center=(col_idx in center_cols),
                    row_idx=row_idx,
                )

        # Grand total row
        total_row = len(estimate_rows) + 2
        ws.cell(row=total_row, column=1, value="")
        ws.cell(row=total_row, column=2, value="합계")
        ws.cell(row=total_row, column=3, value="")
        ws.cell(row=total_row, column=4, value="")
        ws.cell(row=total_row, column=5, value="")
        total_cell = ws.cell(row=total_row, column=6, value=grand_total)
        total_cell.number_format = _NUMBER_FORMAT
        total_cell.alignment = _DATA_ALIGNMENT_RIGHT
        ws.cell(row=total_row, column=7, value="")
        _apply_total_row(ws, total_row, len(headers))

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 3: 세부공정내역 (Process Detail)
    # ------------------------------------------------------------------

    def _build_process_detail(self, wb: Workbook, processes: list[dict]) -> None:
        ws = wb.create_sheet(title="세부공정내역")

        headers = [
            "No", "대공정", "세부공정", "자재", "수량",
            "단위", "자재비", "투입인력", "소요일수", "노무비", "합계",
        ]
        _apply_header_row(ws, row=1, headers=headers)

        number_cols = {5, 7, 9, 10, 11}  # quantity, material_cost, labor_days, labor_cost, total_cost
        center_cols = {1, 6}              # No, unit

        for row_idx, proc in enumerate(processes):
            r = row_idx + 2
            values = [
                proc.get("no", row_idx + 1),
                proc.get("process", ""),
                proc.get("sub_process", ""),
                proc.get("material", ""),
                proc.get("quantity", 0),
                proc.get("unit", ""),
                proc.get("material_cost", 0),
                proc.get("labor", ""),
                proc.get("labor_days", 0),
                proc.get("labor_cost", 0),
                proc.get("total_cost", 0),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                _apply_data_cell(
                    cell,
                    is_number=(col_idx in number_cols),
                    is_center=(col_idx in center_cols),
                    row_idx=row_idx,
                )

        # Totals row
        total_row = len(processes) + 2
        total_material = sum(p.get("material_cost", 0) for p in processes)
        total_labor = sum(p.get("labor_cost", 0) for p in processes)
        total_all = sum(p.get("total_cost", 0) for p in processes)

        ws.cell(row=total_row, column=1, value="")
        ws.cell(row=total_row, column=2, value="합계")
        for col_idx in (3, 4, 5, 6, 8, 9):
            ws.cell(row=total_row, column=col_idx, value="")

        mat_cell = ws.cell(row=total_row, column=7, value=total_material)
        mat_cell.number_format = _NUMBER_FORMAT
        mat_cell.alignment = _DATA_ALIGNMENT_RIGHT

        lab_cell = ws.cell(row=total_row, column=10, value=total_labor)
        lab_cell.number_format = _NUMBER_FORMAT
        lab_cell.alignment = _DATA_ALIGNMENT_RIGHT

        all_cell = ws.cell(row=total_row, column=11, value=total_all)
        all_cell.number_format = _NUMBER_FORMAT
        all_cell.alignment = _DATA_ALIGNMENT_RIGHT

        _apply_total_row(ws, total_row, len(headers))
        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 4: 공정별요약 (Process Summary)
    # ------------------------------------------------------------------

    def _build_process_summary(self, wb: Workbook, summary: dict) -> None:
        ws = wb.create_sheet(title="공정별요약")

        headers = ["공정명", "세부공정수", "자재비", "노무비", "합계", "예상소요일"]
        _apply_header_row(ws, row=1, headers=headers)

        number_cols = {2, 3, 4, 5, 6}
        summary_values = list(summary.values())

        for row_idx, s in enumerate(summary_values):
            r = row_idx + 2
            values = [
                s.get("process", ""),
                s.get("sub_count", 0),
                s.get("material_cost", 0),
                s.get("labor_cost", 0),
                s.get("total_cost", 0),
                s.get("total_days", 0),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                _apply_data_cell(
                    cell,
                    is_number=(col_idx in number_cols),
                    is_center=(col_idx == 1),
                    row_idx=row_idx,
                )

        # Totals row
        total_row = len(summary_values) + 2
        total_mat = sum(s.get("material_cost", 0) for s in summary_values)
        total_lab = sum(s.get("labor_cost", 0) for s in summary_values)
        total_all = sum(s.get("total_cost", 0) for s in summary_values)
        total_days = sum(s.get("total_days", 0) for s in summary_values)

        ws.cell(row=total_row, column=1, value="합계")
        ws.cell(row=total_row, column=2, value="")

        for col_idx, val in [(3, total_mat), (4, total_lab), (5, total_all), (6, round(total_days, 1))]:
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.number_format = _NUMBER_FORMAT
            cell.alignment = _DATA_ALIGNMENT_RIGHT

        _apply_total_row(ws, total_row, len(headers))
        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 5: 남은작업 (Remaining Work)
    # ------------------------------------------------------------------

    def _build_remaining_work(
        self,
        wb: Workbook,
        processes: list[dict],
        remaining_work: list[dict] | None,
    ) -> None:
        ws = wb.create_sheet(title="남은작업")

        headers = [
            "No", "공정", "세부공정", "자재", "전체수량",
            "완료수량", "잔여수량", "단위", "진행률(%)", "상태", "비고",
        ]
        _apply_header_row(ws, row=1, headers=headers)

        # Build row data: use remaining_work if provided, otherwise derive
        # from processes with defaults (completed_qty=0, status="대기")
        work_items = self._prepare_remaining_items(processes, remaining_work)

        number_cols = {5, 6, 7}   # 전체수량, 완료수량, 잔여수량
        center_cols = {1, 8, 10}  # No, 단위, 상태

        data_start_row = 2
        for row_idx, item in enumerate(work_items):
            r = data_start_row + row_idx
            no = row_idx + 1

            # Static values
            ws.cell(row=r, column=1, value=no)
            ws.cell(row=r, column=2, value=item.get("process", ""))
            ws.cell(row=r, column=3, value=item.get("sub_process", ""))
            ws.cell(row=r, column=4, value=item.get("material", ""))
            ws.cell(row=r, column=5, value=item.get("quantity", 0))
            ws.cell(row=r, column=6, value=item.get("completed_qty", 0))
            ws.cell(row=r, column=8, value=item.get("unit", ""))

            # 잔여수량 = 전체수량 - 완료수량 (Excel formula)
            qty_col = get_column_letter(5)   # E
            done_col = get_column_letter(6)  # F
            ws.cell(row=r, column=7, value=f"={qty_col}{r}-{done_col}{r}")

            # 진행률 = IFERROR(완료수량/전체수량, 0)
            ws.cell(
                row=r,
                column=9,
                value=f'=IFERROR({done_col}{r}/{qty_col}{r},0)',
            )
            ws.cell(row=r, column=9).number_format = "0.0%"

            # 상태
            status = item.get("status", "대기")
            ws.cell(row=r, column=10, value=status)

            # 비고
            ws.cell(row=r, column=11, value=item.get("note", ""))

            # Apply cell styling
            for col_idx in range(1, 12):
                cell = ws.cell(row=r, column=col_idx)
                _apply_data_cell(
                    cell,
                    is_number=(col_idx in number_cols),
                    is_center=(col_idx in center_cols),
                    row_idx=row_idx,
                )

            # Override 진행률 format (column 9 is treated special)
            progress_cell = ws.cell(row=r, column=9)
            progress_cell.number_format = "0.0%"
            progress_cell.alignment = _DATA_ALIGNMENT_CENTER

            # Apply status fill color
            status_cell = ws.cell(row=r, column=10)
            if status in _STATUS_FILLS:
                status_cell.fill = _STATUS_FILLS[status]

        # --- Data validation: dropdown for 상태 column ---
        if work_items:
            last_data_row = data_start_row + len(work_items) - 1
            status_col_letter = get_column_letter(10)  # J
            dv = DataValidation(
                type="list",
                formula1='"대기,진행중,완료"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error = "대기, 진행중, 완료 중 선택하세요."
            dv.errorTitle = "유효하지 않은 상태"
            dv.prompt = "상태를 선택하세요."
            dv.promptTitle = "상태"
            dv.add(f"{status_col_letter}{data_start_row}:{status_col_letter}{last_data_row}")
            ws.add_data_validation(dv)

            # --- DataBar conditional formatting on 진행률 column ---
            progress_col_letter = get_column_letter(9)  # I
            rule = DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color="3B82F6",
            )
            ws.conditional_formatting.add(
                f"{progress_col_letter}{data_start_row}:{progress_col_letter}{last_data_row}",
                rule,
            )

        _auto_width(ws)

    @staticmethod
    def _prepare_remaining_items(
        processes: list[dict],
        remaining_work: list[dict] | None,
    ) -> list[dict]:
        """Merge process data with optional remaining-work overrides.

        When *remaining_work* is ``None``, every process row becomes a
        remaining-work item with ``completed_qty=0`` and ``status="대기"``.

        When *remaining_work* is provided, it is used directly; any missing
        fields are filled with sensible defaults.
        """
        if remaining_work is not None:
            items: list[dict] = []
            for rw in remaining_work:
                items.append({
                    "process": rw.get("process", ""),
                    "sub_process": rw.get("sub_process", ""),
                    "material": rw.get("material", ""),
                    "quantity": rw.get("quantity", 0),
                    "completed_qty": rw.get("completed_qty", 0),
                    "unit": rw.get("unit", ""),
                    "status": rw.get("status", "대기"),
                    "note": rw.get("note", ""),
                })
            return items

        # Derive from processes
        items = []
        for p in processes:
            items.append({
                "process": p.get("process", ""),
                "sub_process": p.get("sub_process", ""),
                "material": p.get("material", ""),
                "quantity": p.get("quantity", 0),
                "completed_qty": 0,
                "unit": p.get("unit", ""),
                "status": "대기",
                "note": "",
            })
        return items

    # ------------------------------------------------------------------
    # Sheet 6: 대시보드 (Dashboard)
    # ------------------------------------------------------------------

    def _build_dashboard(self, wb: Workbook, summary: dict) -> None:
        ws = wb.create_sheet(title="대시보드")

        summary_values = list(summary.values())
        if not summary_values:
            ws.cell(row=1, column=1, value="데이터 없음")
            return

        # --- Data table for charts (top-left area) ---
        # Table 1: 자재비 vs 노무비 (for PieChart)
        ws.cell(row=1, column=1, value="비용 구분")
        ws.cell(row=1, column=2, value="금액")
        ws["A1"].font = _HEADER_FONT
        ws["A1"].fill = _HEADER_FILL
        ws["A1"].alignment = _HEADER_ALIGNMENT
        ws["A1"].border = _THIN_BORDER
        ws["B1"].font = _HEADER_FONT
        ws["B1"].fill = _HEADER_FILL
        ws["B1"].alignment = _HEADER_ALIGNMENT
        ws["B1"].border = _THIN_BORDER

        total_material = sum(s.get("material_cost", 0) for s in summary_values)
        total_labor = sum(s.get("labor_cost", 0) for s in summary_values)

        ws.cell(row=2, column=1, value="자재비")
        ws.cell(row=2, column=2, value=total_material)
        ws["B2"].number_format = _NUMBER_FORMAT
        ws.cell(row=3, column=1, value="노무비")
        ws.cell(row=3, column=2, value=total_labor)
        ws["B3"].number_format = _NUMBER_FORMAT

        for r in (2, 3):
            for c in (1, 2):
                cell = ws.cell(row=r, column=c)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER

        # PieChart: 자재비 vs 노무비
        pie = PieChart()
        pie.title = "자재비 vs 노무비 비율"
        pie.style = 10
        pie.width = 16
        pie.height = 12

        labels = Reference(ws, min_col=1, min_row=2, max_row=3)
        data = Reference(ws, min_col=2, min_row=1, max_row=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        # Color slices
        if pie.series:
            series = pie.series[0]
            from openpyxl.chart.series import DataPoint
            # Material cost slice - blue
            pt0 = DataPoint(idx=0)
            pt0.graphicalProperties.solidFill = "3B82F6"
            series.data_points.append(pt0)
            # Labor cost slice - orange
            pt1 = DataPoint(idx=1)
            pt1.graphicalProperties.solidFill = "F59E0B"
            series.data_points.append(pt1)

        ws.add_chart(pie, "D1")

        # --- Table 2: 공정별 비용 (for BarChart) ---
        table2_start_row = 6
        ws.cell(row=table2_start_row, column=1, value="공정명")
        ws.cell(row=table2_start_row, column=2, value="자재비")
        ws.cell(row=table2_start_row, column=3, value="노무비")
        ws.cell(row=table2_start_row, column=4, value="합계")

        for c in range(1, 5):
            cell = ws.cell(row=table2_start_row, column=c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

        for idx, s in enumerate(summary_values):
            r = table2_start_row + 1 + idx
            ws.cell(row=r, column=1, value=s.get("process", ""))
            ws.cell(row=r, column=2, value=s.get("material_cost", 0))
            ws.cell(row=r, column=3, value=s.get("labor_cost", 0))
            ws.cell(row=r, column=4, value=s.get("total_cost", 0))

            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                if c >= 2:
                    cell.number_format = _NUMBER_FORMAT
                    cell.alignment = _DATA_ALIGNMENT_RIGHT

        # BarChart: 공정별 비용
        bar_chart_start = table2_start_row
        bar_chart_end = table2_start_row + len(summary_values)

        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.title = "공정별 비용"
        bar.y_axis.title = "금액 (원)"
        bar.x_axis.title = "공정"
        bar.style = 10
        bar.width = 22
        bar.height = 14

        categories = Reference(ws, min_col=1, min_row=bar_chart_start + 1, max_row=bar_chart_end)
        material_data = Reference(ws, min_col=2, min_row=bar_chart_start, max_row=bar_chart_end)
        labor_data = Reference(ws, min_col=3, min_row=bar_chart_start, max_row=bar_chart_end)

        bar.add_data(material_data, titles_from_data=True)
        bar.add_data(labor_data, titles_from_data=True)
        bar.set_categories(categories)

        # Color the series
        if len(bar.series) >= 2:
            bar.series[0].graphicalProperties.solidFill = "3B82F6"  # Material - blue
            bar.series[1].graphicalProperties.solidFill = "F59E0B"  # Labor - orange

        # Place below the pie chart
        anchor_row = max(bar_chart_end + 2, 18)
        ws.add_chart(bar, f"D{anchor_row}")

        _auto_width(ws)
